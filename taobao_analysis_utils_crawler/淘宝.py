#!/usr/bin/python3
# encoding: utf-8 
# @Time    : 2018/9/27 20:56 
# @author  : zza
# @Email   : 740713651@qq.com
import json
import os
import time
from urllib.parse import urlencode

from openpyxl import load_workbook, Workbook

import requests


def main():
    with open("淘宝cookies.txt", "r", encoding="utf8") as f:
        header = f.readlines()
    with open("关键词.txt", "r", encoding="utf8") as f:
        keywords = f.readlines()
    headers = {}
    for i in header:
        if i.strip() is "":
            continue
        x, y = i.strip().split(": ")
        headers[x] = y
    url = headers.pop("Request URL").strip()
    start = url.find("keyword")
    end = start + url[start:].find("&")
    url = url.replace(url[start:  end], "{}")
    data_table = {}

    xls_file = "淘宝_{}.xls".format(time.strftime("%Y-%m-%d", time.localtime()))
    if os.path.exists(xls_file):
        os.remove(xls_file)
    wb = Workbook()

    def get_items(keyword):
        url1 = url.format(str(urlencode({"keyword": keyword}))).replace('%EF%BB%BF', '')
        rep = requests.get(url1, headers=headers)
        datas = json.loads(rep.text)
        datas = datas["content"]["data"]
        print("over")
        data_table[keyword] = datas

    def save_2_doc(key):
        ws = wb.create_sheet(key)
        keys = list(data_table[key][0].keys())
        ws.append(keys + ["筛选条件"])
        for i in data_table[key]:
            if i.get("suv") * 7 - i.get("onlineGoodsCnt") > 0 \
                    and i.get("payConvRate") is not None \
                    and i.get("payConvRate") > 0.05:
                f = [i.get(j) for j in keys] + [i.get("suv") * 7 - i.get("onlineGoodsCnt")]
                ws.append(f)
        print("{} ok".format(key))

    for i in keywords:
        if i.strip() == "":
            continue
        print(i.strip(), end="  ")
        get_items(i.strip())
        time.sleep(2)
    print("*" * 30)
    print("开始保存文件")
    for i in data_table:
        save_2_doc(i)

    wb.save(xls_file)


if __name__ == '__main__':
    try:
        print("开始爬虫")
        print("*" * 30)
        main()
        print("完成")
        print("*" * 30)
    except Exception as err:
        print("*" * 30)
        print("系统错误")
        print("错误原因:")
        print(err)
        input()
